import openpyxl

workbook = openpyxl.load_workbook("parameters.xlsx")
worksheet = workbook.active

# you can also select the worksheet by name
worksheet_by_name = workbook["Tabelle1"]
assert worksheet == worksheet_by_name


# We now want to search for the cell that contans the parameter "P2"
# and print the value its left neighbor:
for row in worksheet:
    for cell in row:
        if cell.value == "P2":
            i, j = cell.row, cell.column
            print("Found P2 at cell:", i, j)
            print("Cell left of P2 has value:", worksheet.cell(i, j+1).value)


# To modify cells, just assign values,
# either by cell name ...
worksheet["B11"] = "P_new"
worksheet["C11"] = 17.
# ... or by indices. Note the one-based indexing ...
assert worksheet.cell(11, 3).value == 17. # C-->3
worksheet.cell(11, 3).value = 6174.

# ... and save the workbook
workbook.save("parameters_modified.xlsx")


# You can now get creative and abstract functions like ...
import re
def row_col_of(search_expression):
    pattern = re.compile(search_expression)
    result = []
    for row in worksheet:
        for cell in row:
            if pattern.search(str(cell.value)):
                result.append(((cell.row, cell.column), cell.value))
    return result

# ... to not only search for exact strings
print(row_col_of("P2"))

# ... but also t
print(row_col_of("(P.*)"))
